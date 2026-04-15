#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""parse_excel.py - 간단 버전 (openpyxl만 사용)"""
import sys
import json
import re

try:
    import openpyxl
    from datetime import datetime
except ImportError:
    print(json.dumps({
        'error': 'openpyxl 라이브러리가 설치되지 않았습니다',
        'count': 0,
        'data': []
    }, ensure_ascii=False))
    sys.exit(1)

def s(v):
    """값을 안전한 문자열로 변환"""
    if v is None or v == '':
        return ''
    return str(v).strip()

def to_int(v):
    """숫자 변환"""
    try:
        cleaned = str(v).replace(',', '').replace('\t', '').strip()
        return int(float(cleaned))
    except:
        return 0

def parse_date(v):
    """날짜 변환"""
    if v is None:
        return ''
    
    # datetime 객체인 경우
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    
    raw = str(v).strip()
    if not raw or raw in ('None', ''):
        return ''
    
    # YYYY-MM-DD 형식
    m = re.search(r'(\d{4})[-/.\\\\ ](\d{1,2})[-/.\\\\ ](\d{1,2})', raw)
    if m:
        year = m.group(1)
        month = m.group(2).zfill(2)
        day = m.group(3).zfill(2)
        return "{0}-{1}-{2}".format(year, month, day)
    
    return raw[:10]

def parse_members(path):
    """회원정보 파싱"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as e:
        raise Exception("엑셀 파일 읽기 실패: {0}".format(str(e)))
    
    # 헤더 읽기
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[idx] = str(cell.value).strip()
    
    rows = []
    for row_cells in ws.iter_rows(min_row=2):
        row_dict = {}
        for idx, cell in enumerate(row_cells, start=1):
            header = headers.get(idx)
            if header:
                row_dict[header] = cell.value
        
        member_no = s(row_dict.get('회원번호'))
        if not member_no:
            continue
        
        active = s(row_dict.get('탈퇴', 'X'))
        
        rows.append({
            'member_no':      member_no,
            'name':           s(row_dict.get('이름')),
            'login_id':       s(row_dict.get('ID')),
            'serial':         s(row_dict.get('serial')),
            'nationality':    s(row_dict.get('국적')) or '내국인',
            'is_active':      0 if active == 'O' else 1,
            'is_approved':    1 if s(row_dict.get('승인여부')) == 'O' else 0,
            'join_date':      parse_date(row_dict.get('가입일')),
            'center':         s(row_dict.get('센터')),
            'grade':          s(row_dict.get('등급')) or '회원',
            'referrer_no':    s(row_dict.get('추천인번호')),
            'referrer_name':  s(row_dict.get('추천인명')),
            'referrer_id':    s(row_dict.get('추천인ID')),
            'sponsor_no':     s(row_dict.get('후원인번호')),
            'sponsor_name':   s(row_dict.get('후원인명')),
            'sponsor_id':     s(row_dict.get('후원인ID')),
            'position':       s(row_dict.get('위치')),
            'phone':          s(row_dict.get('휴대폰')),
            'gender':         s(row_dict.get('성별')),
            'postal_code':    s(row_dict.get('우편번호')),
            'address1':       s(row_dict.get('주소1')),
            'address2':       s(row_dict.get('주소2')),
            'bank_name':      s(row_dict.get('은행명')),
            'account_holder': s(row_dict.get('예금주')),
            'account_no':     s(row_dict.get('계좌번호')),
            'email':          s(row_dict.get('이메일')),
            'memo':           s(row_dict.get('비고')),
        })
    
    if not rows:
        raise Exception('유효한 회원 데이터가 없습니다')
    
    return rows

def parse_sales(path):
    """매출 데이터 파싱"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
    except Exception as e:
        raise Exception("엑셀 파일 읽기 실패: {0}".format(str(e)))
    
    # 헤더 읽기
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[idx] = str(cell.value).strip()
    
    rows = []
    for row_cells in ws.iter_rows(min_row=2):
        row_dict = {}
        for idx, cell in enumerate(row_cells, start=1):
            header = headers.get(idx)
            if header:
                row_dict[header] = cell.value
        
        name = s(row_dict.get('이름'))
        if not name:
            continue
        
        rows.append({
            'order_type':    s(row_dict.get('주문종류', '1')).replace('\t', '').strip() or '1',
            'order_date':    parse_date(row_dict.get('주문일')),
            'member_name':   name,
            'ddm_id':        s(row_dict.get('DDM ID')),
            'phone':         s(row_dict.get('연락처')),
            'product_name':  s(row_dict.get('상품명')),
            'amount':        to_int(row_dict.get('주문금액', 0)),
            'pv':            to_int(row_dict.get('BV', 0)),
            'pv2':           to_int(row_dict.get('BV2', 0)),
            'memo':          s(row_dict.get('비고')),
            'mall_no':       s(row_dict.get('쇼핑몰번호')),
        })
    
    if not rows:
        raise Exception('유효한 매출 데이터가 없습니다')
    
    return rows

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({
            'error': '사용법: parse_excel.py <파일경로> <타입>',
            'count': 0,
            'data': []
        }, ensure_ascii=False))
        sys.exit(1)
    
    path = sys.argv[1]
    typ = sys.argv[2]
    
    try:
        if typ == 'members':
            rows = parse_members(path)
        elif typ == 'sales':
            rows = parse_sales(path)
        else:
            raise Exception('알 수 없는 타입: {0}'.format(typ))
        
        print(json.dumps({
            'count': len(rows),
            'data': rows
        }, ensure_ascii=False))
        
    except Exception as e:
        print(json.dumps({
            'error': str(e),
            'count': 0,
            'data': []
        }, ensure_ascii=False))
        sys.exit(1)
